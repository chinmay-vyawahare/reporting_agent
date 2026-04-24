"""
Single-chart → Excel (.xlsx) export.

Takes an API-shaped chart dict (same format the canvas / template / SSE
endpoints produce) and writes a tabular .xlsx workbook that captures the
chart's underlying data, shaped for whichever chart type it is:

  * Cartesian (column / bar / line / area / spline / areaspline / scatter):
      One "Data" sheet — first column = category, then one column per series.
      If the chart has N series they appear as N side-by-side value columns.

  * Pie / donut:
      One "Data" sheet — three columns: Name · Value · %.

A small "About" sheet appears at the front with the chart title, subtitle,
description, and insight so the spreadsheet is self-contained (reader can
see the chart's headline without a separate PDF).
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_CARTESIAN_TYPES = ("column", "bar", "line", "area", "spline", "areaspline", "scatter")
_PIE_TYPES = ("pie", "donut")


def _title_text(v: Any) -> str:
    if isinstance(v, dict):
        return v.get("text") or ""
    return v or ""


def _chart_type(chart: dict[str, Any]) -> str:
    inner = chart.get("chart") or {}
    return str((inner.get("type") if isinstance(inner, dict) else None) or "").lower()


def _apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_title_style(cell) -> None:
    cell.font = Font(bold=True, size=14, color="0F172A")


def _apply_meta_label(cell) -> None:
    cell.font = Font(bold=True, color="1F2937")
    cell.alignment = Alignment(vertical="top")


def _autosize_columns(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for c in col:
            v = "" if c.value is None else str(c.value)
            for line in v.split("\n"):
                length = max(length, len(line))
        ws.column_dimensions[letter].width = min(max_width, max(12, length + 2))


def _write_about_sheet(wb: Workbook, chart: dict[str, Any]) -> None:
    """Front sheet with human-readable context — title, subtitle, description,
    insight. Keeps the workbook self-explanatory when opened standalone."""
    ws = wb.create_sheet(title="About", index=0)

    ws["A1"] = _title_text(chart.get("title"))
    _apply_title_style(ws["A1"])
    ws.merge_cells("A1:D1")

    row = 3
    sub = _title_text(chart.get("subtitle"))
    desc = (chart.get("description") or "").strip()
    insight = (chart.get("insight") or "").strip()

    for label, value in (
        ("Subtitle",    sub),
        ("Description", desc),
        ("Insight",     insight),
        ("Chart ID",    chart.get("chart_id") or ""),
        ("Chart type",  _chart_type(chart)),
    ):
        if not value:
            continue
        ws.cell(row=row, column=1, value=label)
        _apply_meta_label(ws.cell(row=row, column=1))
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        if label == "Insight":
            ws.row_dimensions[row].height = 56
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


def _write_cartesian_data(wb: Workbook, chart: dict[str, Any]) -> None:
    """Write a wide-format table: first column = x category, then one
    column per series."""
    ws = wb.create_sheet(title="Data")

    xaxis = chart.get("xAxis") or {}
    cats  = list(xaxis.get("categories") or [])
    series = [s for s in (chart.get("series") or []) if isinstance(s, dict)]

    x_title = _title_text(xaxis.get("title")) or "Category"
    y_title = _title_text((chart.get("yAxis") or {}).get("title"))

    # Header row: [x-axis title, series1 name, series2 name, …]
    ws.cell(row=1, column=1, value=x_title)
    _apply_header_style(ws.cell(row=1, column=1))
    for i, s in enumerate(series):
        header = s.get("name") or f"Series {i+1}"
        if y_title and len(series) == 1:
            header = f"{header} ({y_title})"
        c = ws.cell(row=1, column=i + 2, value=header)
        _apply_header_style(c)

    # Body: one row per category.
    for r, cat in enumerate(cats, start=2):
        ws.cell(row=r, column=1, value=cat)
        for si, s in enumerate(series):
            data = s.get("data") or []
            v = data[r - 2] if r - 2 < len(data) else None
            # Coerce numeric-ish strings back to numbers so Excel treats
            # them as numbers (enables formulas, sum, etc.)
            if isinstance(v, str):
                try: v = float(v)
                except Exception: pass
            ws.cell(row=r, column=si + 2, value=v)

    ws.freeze_panes = "A2"    # keep header row visible when scrolling
    _autosize_columns(ws)


def _write_pie_data(wb: Workbook, chart: dict[str, Any]) -> None:
    """Pie / donut: Name · Value · % of total."""
    ws = wb.create_sheet(title="Data")

    series = chart.get("series") or []
    slices: list[dict] = []
    if series and isinstance(series[0], dict):
        slices = [s for s in (series[0].get("data") or []) if isinstance(s, dict)]

    # Header
    headers = ["Name", "Value", "% of Total"]
    for i, h in enumerate(headers, start=1):
        _apply_header_style(ws.cell(row=1, column=i, value=h))

    total = sum((s.get("y") or 0) for s in slices) or 1
    for r, s in enumerate(slices, start=2):
        y = s.get("y") or 0
        try: y = float(y)
        except Exception: pass
        ws.cell(row=r, column=1, value=s.get("name") or f"Slice {r-1}")
        ws.cell(row=r, column=2, value=y)
        pct_cell = ws.cell(row=r, column=3, value=(y / total) if isinstance(y, (int, float)) else None)
        pct_cell.number_format = "0.0%"

    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def render_chart_xlsx(chart: dict[str, Any]) -> bytes:
    """Return the chart as xlsx bytes.

    Structure:
        About   — title / subtitle / description / insight / chart_id
        Data    — the tabular form of the chart (per-type shape above)
    """
    wb = Workbook()
    # Workbook starts with a default "Sheet" — we create About + Data and
    # drop the default.
    default = wb.active
    _write_about_sheet(wb, chart)

    ctype = _chart_type(chart)
    if ctype in _CARTESIAN_TYPES:
        _write_cartesian_data(wb, chart)
    elif ctype in _PIE_TYPES:
        _write_pie_data(wb, chart)
    else:
        ws = wb.create_sheet(title="Data")
        ws["A1"] = f"Unsupported chart type: {ctype!r}"

    wb.remove(default)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
